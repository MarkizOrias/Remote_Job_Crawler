"""Match scraped job listings against profile.json and rank by fit."""

import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def load_profile() -> dict:
    return json.loads(Path("profile.json").read_text(encoding="utf-8"))


def load_scraped() -> list[dict]:
    path = Path("data/scraped_roles.json")
    if not path.exists():
        legacy = Path("data/scraped_roles.jsonl")
        raw: list[dict] = []
        with legacy.open("r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line:
                    raw.append(json.loads(line))
    else:
        raw = json.loads(path.read_text(encoding="utf-8"))
    return [r for r in raw if r.get("text") and not r.get("error")]


# Keywords that indicate a link points to an individual job posting
JOB_LINK_RE = re.compile(
    r"(analyst|specialist|engineer|manager|lead|director|coordinator|associate|"
    r"officer|developer|architect|admin|accountant|operation|finance|data|"
    r"reconcil|settle|trade|report|automat|python|support)",
    re.IGNORECASE,
)

# URL patterns for known ATS job detail pages
ATS_JOB_RE = re.compile(
    r"(greenhouse\.io/.+/jobs/|lever\.co/.+/|ashbyhq\.com/.+/|"
    r"myworkdayjobs\.com/.+/job/|workable\.com/.+/j/|"
    r"bamboohr\.com/careers/|/jobs?/\d|/positions?/\d|/careers/.+/\d)",
    re.IGNORECASE,
)


def extract_job_links(links: list[dict], text_lower: str, profile: dict) -> list[dict]:
    """Filter page links down to those likely pointing to relevant job postings."""
    prefs = profile["preferences"]
    relevant = []

    for link in links:
        href = link.get("href", "")
        text = link.get("text", "")
        combined = (text + " " + href).lower()

        # Skip generic nav/footer links
        if not text or len(text) < 3:
            continue
        if any(skip in combined for skip in ["login", "sign in", "privacy", "cookie", "terms of"]):
            continue

        # Check if link text or URL matches job-related patterns
        is_job = bool(JOB_LINK_RE.search(text)) or bool(ATS_JOB_RE.search(href))
        if not is_job:
            continue

        # Check for exclude keywords
        if any(kw.lower() in combined for kw in prefs["exclude_keywords"]):
            continue

        relevant.append({"title": text.strip(), "url": href})

    return relevant


def score_company(text_lower: str, profile: dict) -> dict:
    """Score a scraped page against the profile. Returns score breakdown."""
    prefs = profile["preferences"]
    cv = profile["cv"]

    scores = {}

    # 1. Preferred keywords (2 points each)
    kw_hits = []
    for kw in prefs["preferred_keywords"]:
        if kw.lower() in text_lower:
            kw_hits.append(kw)
    scores["preferred_keywords"] = kw_hits
    scores["preferred_kw_score"] = len(kw_hits) * 2

    # 2. Target role title matches (5 points each)
    role_hits = []
    for role in prefs["roles"]:
        # Match individual significant words from role titles
        words = [w for w in role.lower().split() if len(w) > 3]
        if any(w in text_lower for w in words):
            role_hits.append(role)
    scores["role_title_hits"] = role_hits
    scores["role_title_score"] = len(role_hits) * 5

    # 3. Operations/finance skill matches (3 points each)
    ops_hits = []
    for skill in cv["skills"]["operations_finance"]:
        if skill.lower() in text_lower:
            ops_hits.append(skill)
    scores["ops_finance_hits"] = ops_hits
    scores["ops_finance_score"] = len(ops_hits) * 3

    # 4. Technical skill matches (2 points each)
    tech_hits = []
    for skill in cv["skills"]["technical"]:
        # Exact word boundary match for short terms like "Python", "CSS"
        pattern = r"\b" + re.escape(skill.lower()) + r"\b"
        if re.search(pattern, text_lower):
            tech_hits.append(skill)
    scores["tech_hits"] = tech_hits
    scores["tech_score"] = len(tech_hits) * 2

    # 5. Methodology matches (2 points each)
    method_hits = []
    for skill in cv["skills"]["methodology"]:
        if skill.lower() in text_lower:
            method_hits.append(skill)
    scores["method_hits"] = method_hits
    scores["method_score"] = len(method_hits) * 2

    # 6. Exclude keyword penalty (-100 each, effectively disqualifies)
    exclude_hits = []
    for kw in prefs["exclude_keywords"]:
        if kw.lower() in text_lower:
            exclude_hits.append(kw)
    scores["exclude_hits"] = exclude_hits
    scores["exclude_penalty"] = len(exclude_hits) * -100

    # Total
    scores["total"] = (
        scores["preferred_kw_score"]
        + scores["role_title_score"]
        + scores["ops_finance_score"]
        + scores["tech_score"]
        + scores["method_score"]
        + scores["exclude_penalty"]
    )

    return scores


def main():
    profile = load_profile()
    scraped = load_scraped()
    exclude_companies = [c.lower() for c in profile["preferences"].get("exclude_companies", [])]

    print(f"Loaded {len(scraped)} successful scrapes. Scoring...\n")

    results = []
    for entry in scraped:
        company = entry["company"]
        if company.lower() in exclude_companies:
            continue

        text_lower = entry["text"].lower()
        scores = score_company(text_lower, profile)

        if scores["total"] > 0:
            job_links = extract_job_links(
                entry.get("links", []), text_lower, profile,
            )
            results.append({
                "company": company,
                "careers_url": entry["careers_url"],
                "final_url": entry["final_url"],
                "job_links": job_links,
                "scores": scores,
            })

    results.sort(key=lambda r: r["scores"]["total"], reverse=True)

    # Print top matches
    print(f"{'Rank':<5} {'Score':<6} {'Company':<35} {'Role Hits'}")
    print("-" * 110)

    for i, r in enumerate(results[:30], 1):
        s = r["scores"]
        role_summary = ", ".join(s["role_title_hits"][:3]) or "-"
        kw_summary = ", ".join(s["preferred_keywords"][:4])
        print(f"{i:<5} {s['total']:<6} {r['company'][:34]:<35} {role_summary}")
        print(f"{'':5} {'':6} Keywords: {kw_summary}")
        if s["ops_finance_hits"]:
            print(f"{'':5} {'':6} Ops/Finance: {', '.join(s['ops_finance_hits'])}")
        if s["tech_hits"]:
            print(f"{'':5} {'':6} Tech: {', '.join(s['tech_hits'])}")
        if s["exclude_hits"]:
            print(f"{'':5} {'':6} ** EXCLUDED: {', '.join(s['exclude_hits'])}")
        print(f"{'':5} {'':6} URL: {r['final_url']}")
        if r.get("job_links"):
            for jl in r["job_links"][:5]:
                print(f"{'':5} {'':6}   -> {jl['title'][:60]}")
                print(f"{'':5} {'':6}      {jl['url'][:120]}")
        print()

    # Save full results to JSON
    out_path = Path("data/matched_roles.json")
    out_path.write_text(
        json.dumps(results, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    print(f"\nFull results ({len(results)} matches) saved to {out_path}")

    xlsx_path = Path("data/matched_roles.xlsx")
    write_excel(results, xlsx_path)
    print(f"Excel workbook saved to {xlsx_path}")


def write_excel(results: list[dict], path: Path) -> None:
    """Write a single-sheet xlsx of every >0-score match.

    Columns are ordered for at-a-glance triage: rank, score, company, then the
    keyword/skill hit columns that explain *why* the score is what it is, then
    the URLs and the discovered job links.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Matches"

    headers = [
        "Rank",
        "Score",
        "Company",
        "Role Title Hits",
        "Preferred Keywords",
        "Ops/Finance Hits",
        "Tech Hits",
        "Methodology Hits",
        "Excluded Keywords",
        "Careers URL",
        "Final URL",
        "Top Job Links",
    ]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.freeze_panes = "A2"

    excluded_fill = PatternFill("solid", fgColor="F8CBAD")
    wrap = Alignment(wrap_text=True, vertical="top")

    for rank, r in enumerate(results, start=1):
        s = r["scores"]
        job_links_cell = "\n".join(
            f"{jl['title'][:80]} — {jl['url']}" for jl in r.get("job_links", [])[:8]
        )
        row = [
            rank,
            s["total"],
            r["company"],
            ", ".join(s["role_title_hits"]),
            ", ".join(s["preferred_keywords"]),
            ", ".join(s["ops_finance_hits"]),
            ", ".join(s["tech_hits"]),
            ", ".join(s["method_hits"]),
            ", ".join(s["exclude_hits"]),
            r["careers_url"],
            r["final_url"],
            job_links_cell,
        ]
        ws.append(row)
        excel_row = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=excel_row, column=col_idx).alignment = wrap
        # Visually flag rows that hit an exclusion keyword.
        if s["exclude_hits"]:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=excel_row, column=col_idx).fill = excluded_fill

    column_widths = [6, 7, 28, 38, 28, 28, 24, 20, 20, 45, 45, 70]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


if __name__ == "__main__":
    main()
