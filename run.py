"""Consolidated pipeline: scrape all sources → match against profile → Excel.

Usage:
    python run.py              # full run (job boards + career page crawler)
    python run.py --fast       # job boards only (skip the 30-min career crawler)
"""

import asyncio
import json
import logging
import re
import sqlite3
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

PROFILE_PATH = Path("profile.json")
JSONL_PATH = Path("data/scraped_roles.jsonl")
EXCEL_OUTPUT = Path("data/matched_jobs.xlsx")

log = logging.getLogger("run")


def setup() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
        handlers=[
            logging.FileHandler("data/scraper.log", encoding="utf-8"),
            logging.StreamHandler(),
        ],
    )
    Path("data").mkdir(exist_ok=True)


# ===== STAGE 1: Scrape all job board sources (new system) =================

def scrape_job_boards() -> None:
    """Run all src/ scrapers that produce structured jobs → SQLite."""
    from src.config import setup_logging
    from src.db.connection import get_connection, init_db
    from src.db.queries import bulk_upsert_jobs, start_scrape_run, finish_scrape_run
    from src.scrapers.remoteok import RemoteOKScraper
    from src.scrapers.remotive import RemotiveScraper
    from src.scrapers.weworkremotely import WeWorkRemotelyScraper
    from src.scrapers.dayonejobs import DayOneJobsScraper

    init_db()

    scrapers = [
        RemoteOKScraper(),
        RemotiveScraper(),
        WeWorkRemotelyScraper(),
        DayOneJobsScraper(),
    ]

    async def run_all():
        for scraper in scrapers:
            conn = get_connection()
            run_id = start_scrape_run(conn, scraper.name)
            try:
                jobs = await scraper.scrape()
                db_dicts = [j.to_db_dict() for j in jobs]
                total, new_count = bulk_upsert_jobs(conn, db_dicts)
                finish_scrape_run(conn, run_id, jobs_found=total, jobs_new=new_count)
                print(f"  [{scraper.name}] {total} jobs, {new_count} new")
            except Exception as exc:
                finish_scrape_run(conn, run_id, status="failed", error_message=str(exc))
                print(f"  [{scraper.name}] FAILED: {exc}")
            finally:
                await scraper.close()
                conn.close()

    asyncio.run(run_all())


# ===== STAGE 2: Crawl company career pages (old system) ===================

def scrape_career_pages() -> None:
    """Run the career page crawler: clone repo → crawl ~850 pages → JSONL."""
    from scraper.parse_repo import parse_companies
    from scraper.crawl import crawl_all

    print("  Parsing company URLs from remoteintech/remote-jobs repo...")
    companies = parse_companies()
    total = len(companies)
    print(f"  Found {total} companies. Crawling career pages...")

    state = {"done": 0, "errors": 0, "last": 0}

    def progress(idx: int, error):
        state["done"] += 1
        if error:
            state["errors"] += 1
        n = state["done"]
        if n - state["last"] >= 50 or n == total:
            pct = int(n / total * 100)
            print(f"  [{n}/{total}] {pct}% — {state['errors']} errors")
            state["last"] = n

    asyncio.run(crawl_all(companies, JSONL_PATH, progress_cb=progress))

    if JSONL_PATH.exists():
        lines = JSONL_PATH.read_text(encoding="utf-8").strip().splitlines()
        errors = sum(1 for l in lines if '"error": null' not in l)
        print(f"  Career pages done: {len(lines)} records, {errors} with errors")


# ===== STAGE 3: Load & score everything ===================================

def load_profile() -> dict:
    return json.loads(PROFILE_PATH.read_text(encoding="utf-8"))


def load_db_jobs() -> list[dict]:
    """Load structured jobs from SQLite (job board scrapers)."""
    from src.db.connection import get_connection
    conn = get_connection()
    rows = conn.execute(
        "SELECT * FROM jobs WHERE is_active = 1"
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def load_career_page_jobs() -> list[dict]:
    """Load career page scrapes from JSONL and extract individual job links."""
    if not JSONL_PATH.exists():
        return []

    job_link_re = re.compile(
        r"(analyst|specialist|engineer|manager|lead|director|coordinator|associate|"
        r"officer|developer|architect|accountant|operation|finance|data|"
        r"reconcil|settle|trade|report|automat|python|support)",
        re.IGNORECASE,
    )

    results = []
    for line in JSONL_PATH.read_text(encoding="utf-8").strip().splitlines():
        try:
            entry = json.loads(line)
        except json.JSONDecodeError:
            continue
        if entry.get("error") or not entry.get("text"):
            continue

        company = entry["company"]
        links = entry.get("links", [])

        for link in links:
            href = link.get("href", "")
            text = link.get("text", "")
            if not text or len(text) < 5 or not href:
                continue
            if any(skip in (text + href).lower() for skip in [
                "login", "sign in", "privacy", "cookie", "terms of",
            ]):
                continue
            if not job_link_re.search(text):
                continue

            results.append({
                "title": text.strip()[:200],
                "company": company,
                "url": href,
                "source": "career_page",
                "location": None,
                "salary_min": None,
                "salary_max": None,
                "description": entry["text"][:2000],
                "tags": "",
            })

        if not links:
            results.append({
                "title": f"Careers at {company}",
                "company": company,
                "url": entry.get("final_url") or entry["careers_url"],
                "source": "career_page",
                "location": None,
                "salary_min": None,
                "salary_max": None,
                "description": entry["text"][:2000],
                "tags": "",
            })

    return results


def score_job(job: dict, profile: dict) -> dict:
    """Score a single job against the profile."""
    prefs = profile["preferences"]
    cv = profile["cv"]

    title_lower = (job.get("title") or "").lower()
    desc_lower = (job.get("description") or "").lower()
    tags_lower = (job.get("tags") or "").lower()
    combined = f"{title_lower} {desc_lower} {tags_lower}"

    role_hits = []
    for role in prefs["roles"]:
        words = [w for w in role.lower().split() if len(w) > 3]
        if any(w in title_lower for w in words):
            role_hits.append(role)

    kw_hits = [kw for kw in prefs["preferred_keywords"] if kw.lower() in combined]
    ops_hits = [s for s in cv["skills"]["operations_finance"] if s.lower() in combined]

    tech_hits = []
    for skill in cv["skills"]["technical"]:
        if re.search(r"\b" + re.escape(skill.lower()) + r"\b", combined):
            tech_hits.append(skill)

    method_hits = [s for s in cv["skills"]["methodology"] if s.lower() in combined]
    exclude_hits = [kw for kw in prefs["exclude_keywords"] if kw.lower() in combined]

    total = (
        len(role_hits) * 5
        + len(kw_hits) * 2
        + len(ops_hits) * 3
        + len(tech_hits) * 2
        + len(method_hits) * 2
        + len(exclude_hits) * -100
    )

    return {
        "total": total,
        "role_hits": role_hits,
        "keyword_hits": kw_hits,
        "ops_hits": ops_hits,
        "tech_hits": tech_hits,
        "method_hits": method_hits,
        "exclude_hits": exclude_hits,
    }


def match_all(jobs: list[dict], profile: dict) -> list[dict]:
    """Score and filter all jobs, return sorted matches."""
    exclude_companies = {c.lower() for c in profile["preferences"].get("exclude_companies", [])}
    seen_urls: set[str] = set()
    results = []

    for job in jobs:
        company = (job.get("company") or "").strip()
        if company.lower() in exclude_companies:
            continue

        url = job.get("url", "")
        if url in seen_urls:
            continue
        seen_urls.add(url)

        scores = score_job(job, profile)
        if scores["total"] <= 0:
            continue

        results.append({"job": job, "scores": scores})

    results.sort(key=lambda r: r["scores"]["total"], reverse=True)
    return results


# ===== STAGE 4: Export to Excel ===========================================

def write_excel(matches: list[dict], output: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Matched Jobs"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

    headers = [
        "Rank", "Score", "Company", "Title", "Location", "Salary",
        "Source", "Role Matches", "Keyword Matches", "Skills Matched",
        "Apply Link",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for i, match in enumerate(matches, 1):
        job = match["job"]
        s = match["scores"]
        row = i + 1

        salary = ""
        if job.get("salary_min") or job.get("salary_max"):
            lo = f"${job['salary_min']:,}" if job.get("salary_min") else "?"
            hi = f"${job['salary_max']:,}" if job.get("salary_max") else "?"
            salary = f"{lo} - {hi}"

        role_str = ", ".join(s["role_hits"]) if s["role_hits"] else ""
        kw_str = ", ".join(s["keyword_hits"]) if s["keyword_hits"] else ""
        all_skills = s["ops_hits"] + s["tech_hits"] + s["method_hits"]
        skills_str = ", ".join(all_skills) if all_skills else ""
        apply_url = job.get("url", "")

        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=s["total"])
        ws.cell(row=row, column=3, value=job.get("company", ""))
        ws.cell(row=row, column=4, value=job.get("title", ""))
        ws.cell(row=row, column=5, value=job.get("location", "") or "")
        ws.cell(row=row, column=6, value=salary)
        ws.cell(row=row, column=7, value=job.get("source", ""))
        ws.cell(row=row, column=8, value=role_str)
        ws.cell(row=row, column=9, value=kw_str)
        ws.cell(row=row, column=10, value=skills_str)

        link_cell = ws.cell(row=row, column=11, value="Apply")
        if apply_url:
            link_cell.hyperlink = apply_url
            link_cell.font = Font(color="0563C1", underline="single")

        if i % 2 == 0:
            shade = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).fill = shade

    widths = [6, 7, 28, 45, 25, 20, 16, 35, 35, 40, 12]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:K{len(matches) + 1}"

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


# ===== MAIN ===============================================================

def main():
    fast_mode = "--fast" in sys.argv
    setup()

    print("=" * 60)
    print("  Remote Job Scraper — Full Pipeline")
    print("=" * 60)

    # Stage 1: Job boards
    print("\n[1/4] Scraping job boards...")
    scrape_job_boards()

    # Stage 2: Career pages (skip in fast mode)
    if not fast_mode:
        print("\n[2/4] Crawling company career pages (this takes ~30 min)...")
        try:
            scrape_career_pages()
        except Exception as exc:
            log.warning("Career page crawler failed: %s", exc)
            print(f"  Career page crawl failed: {exc} — continuing with job board data")
    else:
        print("\n[2/4] Skipped career page crawl (--fast mode)")

    # Stage 3: Score against profile
    print("\n[3/4] Matching jobs against profile...")
    profile = load_profile()

    all_jobs: list[dict] = []

    db_jobs = load_db_jobs()
    print(f"  Loaded {len(db_jobs)} jobs from job boards")
    all_jobs.extend(db_jobs)

    if not fast_mode:
        cp_jobs = load_career_page_jobs()
        print(f"  Loaded {len(cp_jobs)} job links from career pages")
        all_jobs.extend(cp_jobs)

    matches = match_all(all_jobs, profile)
    print(f"  {len(matches)} jobs matched your profile")

    if not matches:
        print("\nNo matching jobs found.")
        return

    # Stage 4: Export
    print(f"\n[4/4] Exporting to {EXCEL_OUTPUT}...")
    write_excel(matches, EXCEL_OUTPUT)

    print(f"\n{'=' * 60}")
    print(f"  Done! {len(matches)} matched jobs -> {EXCEL_OUTPUT}")
    print(f"{'=' * 60}")

    print(f"\nTop 15:")
    print(f"{'#':<4} {'Score':<6} {'Source':<14} {'Company':<25} {'Title'}")
    print("-" * 100)
    for i, m in enumerate(matches[:15], 1):
        j = m["job"]
        print(f"{i:<4} {m['scores']['total']:<6} {j['source'][:13]:<14} {j['company'][:24]:<25} {j['title'][:50]}")


if __name__ == "__main__":
    main()
