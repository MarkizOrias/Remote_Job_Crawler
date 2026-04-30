"""Match scraped jobs against profile.json and export to Excel with apply links."""

import json
import re
import sqlite3
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.db.connection import get_connection

PROFILE_PATH = Path("profile.json")
OUTPUT_PATH = Path("data/matched_jobs.xlsx")


def load_profile() -> dict:
    return json.loads(PROFILE_PATH.read_text(encoding="utf-8"))


def load_jobs(conn: sqlite3.Connection) -> list[dict]:
    rows = conn.execute(
        "SELECT * FROM jobs WHERE is_active = 1 AND (description IS NOT NULL OR title IS NOT NULL)"
    ).fetchall()
    return [dict(r) for r in rows]


def score_job(job: dict, profile: dict) -> dict:
    """Score a single job against the profile. Returns score breakdown."""
    prefs = profile["preferences"]
    cv = profile["cv"]

    title_lower = (job.get("title") or "").lower()
    desc_lower = (job.get("description") or "").lower()
    tags_lower = (job.get("tags") or "").lower()
    combined = f"{title_lower} {desc_lower} {tags_lower}"

    scores: dict = {}

    # 1. Role title match (5 pts each) — match against job title
    role_hits = []
    for role in prefs["roles"]:
        words = [w for w in role.lower().split() if len(w) > 3]
        if any(w in title_lower for w in words):
            role_hits.append(role)
    scores["role_hits"] = role_hits
    scores["role_score"] = len(role_hits) * 5

    # 2. Preferred keywords (2 pts each)
    kw_hits = [kw for kw in prefs["preferred_keywords"] if kw.lower() in combined]
    scores["keyword_hits"] = kw_hits
    scores["keyword_score"] = len(kw_hits) * 2

    # 3. Operations/finance skills (3 pts each)
    ops_hits = [s for s in cv["skills"]["operations_finance"] if s.lower() in combined]
    scores["ops_hits"] = ops_hits
    scores["ops_score"] = len(ops_hits) * 3

    # 4. Technical skills (2 pts each) — word boundary match
    tech_hits = []
    for skill in cv["skills"]["technical"]:
        pattern = r"\b" + re.escape(skill.lower()) + r"\b"
        if re.search(pattern, combined):
            tech_hits.append(skill)
    scores["tech_hits"] = tech_hits
    scores["tech_score"] = len(tech_hits) * 2

    # 5. Methodology (2 pts each)
    method_hits = [s for s in cv["skills"]["methodology"] if s.lower() in combined]
    scores["method_hits"] = method_hits
    scores["method_score"] = len(method_hits) * 2

    # 6. Exclude keywords (-100 each)
    exclude_hits = [kw for kw in prefs["exclude_keywords"] if kw.lower() in combined]
    scores["exclude_hits"] = exclude_hits
    scores["exclude_penalty"] = len(exclude_hits) * -100

    scores["total"] = (
        scores["role_score"]
        + scores["keyword_score"]
        + scores["ops_score"]
        + scores["tech_score"]
        + scores["method_score"]
        + scores["exclude_penalty"]
    )

    return scores


def match_jobs(jobs: list[dict], profile: dict) -> list[dict]:
    """Score all jobs and return sorted matches with positive scores."""
    prefs = profile["preferences"]
    exclude_companies = {c.lower() for c in prefs.get("exclude_companies", [])}

    results = []
    for job in jobs:
        company = (job.get("company") or "").strip()
        if company.lower() in exclude_companies:
            continue

        scores = score_job(job, profile)
        if scores["total"] <= 0:
            continue

        results.append({"job": job, "scores": scores})

    results.sort(key=lambda r: r["scores"]["total"], reverse=True)
    return results


def write_excel(matches: list[dict], output: Path) -> None:
    """Write matched jobs to a styled Excel workbook."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Matched Jobs"

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

    headers = [
        "Rank", "Score", "Company", "Title", "Location", "Salary",
        "Source", "Role Matches", "Keyword Matches", "Skills Matched",
        "Apply Link",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for i, match in enumerate(matches, 1):
        job = match["job"]
        s = match["scores"]
        row = i + 1

        salary = ""
        if job.get("salary_min") or job.get("salary_max"):
            lo = f"${job['salary_min']:,}" if job.get("salary_min") else "?"
            hi = f"${job['salary_max']:,}" if job.get("salary_max") else "?"
            salary = f"{lo} – {hi}"

        role_str = ", ".join(s["role_hits"]) if s["role_hits"] else "—"
        kw_str = ", ".join(s["keyword_hits"]) if s["keyword_hits"] else "—"

        all_skills = s["ops_hits"] + s["tech_hits"] + s["method_hits"]
        skills_str = ", ".join(all_skills) if all_skills else "—"

        apply_url = job.get("url", "")

        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=s["total"])
        ws.cell(row=row, column=3, value=job.get("company", ""))
        ws.cell(row=row, column=4, value=job.get("title", ""))
        ws.cell(row=row, column=5, value=job.get("location", ""))
        ws.cell(row=row, column=6, value=salary)
        ws.cell(row=row, column=7, value=job.get("source", ""))
        ws.cell(row=row, column=8, value=role_str)
        ws.cell(row=row, column=9, value=kw_str)
        ws.cell(row=row, column=10, value=skills_str)

        # Clickable hyperlink
        link_cell = ws.cell(row=row, column=11, value="Apply")
        link_cell.hyperlink = apply_url
        link_cell.font = Font(color="0563C1", underline="single")

        # Alternate row shading
        if i % 2 == 0:
            shade = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = shade

    # Column widths
    widths = [6, 7, 28, 45, 25, 20, 16, 35, 35, 40, 12]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:K{len(matches) + 1}"

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def main():
    profile = load_profile()
    conn = get_connection()
    jobs = load_jobs(conn)
    conn.close()

    print(f"Loaded {len(jobs)} active jobs from database.")

    matches = match_jobs(jobs, profile)
    print(f"Found {len(matches)} matching jobs (score > 0).")

    if not matches:
        print("No matches found.")
        return

    write_excel(matches, OUTPUT_PATH)
    print(f"\nExported to {OUTPUT_PATH}")

    # Print top 10
    print(f"\nTop 10 matches:")
    print(f"{'#':<4} {'Score':<6} {'Company':<30} {'Title':<40} {'Source'}")
    print("-" * 120)
    for i, m in enumerate(matches[:10], 1):
        j = m["job"]
        print(f"{i:<4} {m['scores']['total']:<6} {j['company'][:29]:<30} {j['title'][:39]:<40} {j['source']}")


if __name__ == "__main__":
    main()
